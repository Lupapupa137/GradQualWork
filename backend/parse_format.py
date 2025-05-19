import os
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from io import BytesIO
from datetime import datetime

from backend.database import insert_order, clean_data

from openpyxl.utils import get_column_letter


def apply_auto_width_to_all_sheets(wb):
    for ws in wb.worksheets:
        for col_idx, col in enumerate(ws.iter_cols(min_row=4, max_row=ws.max_row), 1):
            max_length = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=0,
            )

            col_letter = get_column_letter(col_idx)

            if col_letter == "B":
                ws.column_dimensions[col_letter].width = min(max_length + 2, 64)
            elif col_letter == "F":
                ws.column_dimensions[col_letter].width = min(max_length + 2, 7)
            else:
                ws.column_dimensions[col_letter].width = max_length + 2

        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0


def split_by_posts_and_export(
    input_file,
    df,
    rules,
    output_folder,
    manager_name,
    comments,
    internal_number,
    delivery_date,
):
    os.makedirs(output_folder, exist_ok=True)
    post_groups = defaultdict(list)

    wb_input = load_workbook(input_file)
    ws_input = wb_input.active
    order_and_date = ws_input["B15"].value

    parts = order_and_date.split("от")
    order_info = parts[0].strip() 
    order_date_raw = parts[1].strip()
    order_date = datetime.strptime(order_date_raw, "%d.%m.%Y").date()

    for _, row in df.iterrows():
        text = str(row["Товары (работы, услуги)"]).lower()
        for key, post in rules:
            if all(part in text for part in key.split()):
                post_name = f"post_{post}"
                post_groups[post_name].append(row)
                break

    archive_wb = Workbook()
    archive_wb.remove(archive_wb.active)

    post_areas = {}

    for post_name in sorted(
        post_groups.keys(), key=lambda name: int(name.split("_")[1])
    ):
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        rows = post_groups[post_name]
        result_df = pd.DataFrame(rows)
        result_df["№"] = range(1, len(result_df) + 1)
        ws = archive_wb.create_sheet(title=post_name)

        # Заголовки
        for col_idx, col_name in enumerate(result_df.columns, 1):
            cell = ws.cell(row=4, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.border = border

        # Данные
        for row_idx, row in enumerate(result_df.itertuples(index=False), start=5):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True)
                cell.border = border

        # Счёт, менеджер, внутренний, готовность
        ws.cell(row=2, column=1, value=order_and_date).font = Font(size=14, bold=True)
        ws.cell(row=2, column=3, value=manager_name).font = Font(size=14, bold=True)
        ws.cell(row=2, column=6, value=internal_number).font = Font(size=14, bold=True)
        ws.cell(row=2, column=8, value=delivery_date).font = Font(size=14, bold=True)

        # Summary
        summary_row = ws.max_row + 1
        total_qty = result_df["Кол-во"].sum()
        total_area = result_df["S"].sum()

        cell_qty = ws.cell(row=summary_row, column=3, value=total_qty)
        cell_qty.font = Font(size=12, bold=True)
        cell_qty.border = border

        cell_area = ws.cell(row=summary_row, column=6, value=total_area)
        cell_area.font = Font(size=12, bold=True)
        cell_area.border = border

        # Комментарии
        comment_row = ws.max_row + 2
        ws.cell(row=comment_row, column=2, value="Комментарии: " + comments).font = (
            Font(size=16, bold=True)
        )

        # Сохраняем площадь поста
        post_num = int(post_name.split("_")[1])
        post_areas[post_num] = round(total_area, 2)

    apply_auto_width_to_all_sheets(archive_wb)

    archive_path = os.path.join(output_folder, f"{order_and_date}.xlsx")
    archive_wb.save(archive_path)

    output_stream = BytesIO()
    archive_wb.save(output_stream)
    output_stream.seek(0)

    order_record = {
        "order_id": order_info,
        "order_date": str(order_date),
        "total_area": round(sum(post_areas.values()), 2),
    }

    for post_num, area in post_areas.items():
        if post_num <= 10:
            order_record[f"area_post_{post_num}"] = area

    filename = f"{order_and_date}.xlsx"
    print(f"Сохранено: {archive_path}")
    return {filename: output_stream.read()}, order_record
